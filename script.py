import random
import re
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


# random.range(0, 3) : 1~2
# random.choice(stat)

## 공통 함수

def find_cell_by_value(ws, keyword):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if keyword.strip() == str(cell.value).strip():
                    return cell.coordinate
    return None



def shift_cell(cell_ref: str, right=0, down=0) -> str:
    col_letter, row = coordinate_from_string(cell_ref)
    col_index = column_index_from_string(col_letter)

    new_col_index = col_index + right
    new_row = row + down

    if new_col_index < 1 or new_row < 1:
        raise ValueError("엑셀 범위를 벗어났습니다.")

    new_col_letter = get_column_letter(new_col_index)
    return f"{new_col_letter}{new_row}"

def get_shifted_cell_value(ws, base_cell: str, right=0, down=0):
    try:
        target_cell = shift_cell(base_cell, right, down)
        return ws[target_cell].value
    except ValueError as e:
        print(f"에러: {e}")
        return None


def get_offset_between_cells(from_cell: str, to_cell: str) -> tuple[int, int]:
    from_col_letter, from_row = coordinate_from_string(from_cell)
    to_col_letter, to_row = coordinate_from_string(to_cell)

    from_col_index = column_index_from_string(from_col_letter)  # C → 3
    to_col_index = column_index_from_string(to_col_letter)      # AD → 30

    right = to_col_index - from_col_index
    down = to_row - from_row

    return right, down

def roll_dice_expression(expr: str):
    pattern = r"(\d*)d(\d+)([+-]\d+)?"
    match = re.fullmatch(pattern, expr.strip())

    if not match:
        raise ValueError(f"잘못된 표현식입니다: {expr}")

    num_dice = int(match.group(1)) if match.group(1) else 1  # 'd6' → 1d6 처리
    sides = int(match.group(2))
    modifier = int(match.group(3)) if match.group(3) else 0

    rolls = [random.randint(1, sides) for _ in range(num_dice)]
    total = sum(rolls) + modifier
    max_possible = num_dice * sides + modifier

    return total, max_possible, rolls 


###################################################################### CoC 구현

def CoC_dice(bonus: int = 0):
    original_roll = random.randint(1, 100)
    ones = original_roll % 10
    tens = original_roll // 10

    tens_candidates = [tens]

    for _ in range(abs(bonus)):
        roll = random.randint(0, 9)
        tens_candidates.append(roll)

    if bonus < 0:
        final_tens = max(tens_candidates)
    elif bonus > 0:
        final_tens = min(tens_candidates)
    else:
        final_tens = tens

    # 최종 결과
    result = [t * 10 + ones for t in tens_candidates]
    for r in result:
        r = 100 if r == 0 else r
    if final_tens * 10 + ones == 0:
        final = 100
    else:
        final = final_tens * 10 + ones

    message = f"{result} → {final}"
    return final, message
        

def CoC_insane_now():
    number = random.randint(1, 10)
    result = f"🎲 1d10 = {number} | "
    if number == 1:
        result += "광기내용1"
    elif number == 2:
        result += "광기내용2"
    elif number == 3:
        result += "광기내용3"
    elif number == 4:
        result += "광기내용4"
    elif number == 5:
        result += "광기내용5"
    elif number == 6:
        result += "광기내용6"
    elif number == 7:
        result += "광기내용7"
    elif number == 8:
        result += "광기내용8"
    elif number == 9:
        result += "광기내용9"
    elif number == 10:
        result += "광기내용10"
    return result

def CoC_insane_summary():
    number = random.randint(1, 10)
    result = f"🎲 1d10 = {number} | "
    if number == 1:
        result += "광기내용1"
    elif number == 2:
        result += "광기내용2"
    elif number == 3:
        result += "광기내용3"
    elif number == 4:
        result += "광기내용4"
    elif number == 5:
        result += "광기내용5"
    elif number == 6:
        result += "광기내용6"
    elif number == 7:
        result += "광기내용7"
    elif number == 8:
        result += "광기내용8"
    elif number == 9:
        result += "광기내용9"
    elif number == 10:
        result += "광기내용10"
    return result

def CoC_damage(id, skill, modifier, tag):
    script = ""
    path = f'CoC/{id}.xlsx'

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # 또는 wb["시트이름"]

    cell = find_cell_by_value(ws, skill) # 무기 찾기
    print(skill)
    print(cell)

    r1, d1 = get_offset_between_cells('C62', 'Z62')
    damage = get_shifted_cell_value(ws, cell, right=r1, down=d1) #피해 찾기
    print(damage)
    r2, d2 = get_offset_between_cells('C62', 'S62')
    s = get_shifted_cell_value(ws, cell, right=r2, down=d2) # 판정 기능 찾기
    print(s)
    r3, d3 = get_offset_between_cells('C62', 'AD62')
    db = get_shifted_cell_value(ws, cell, right=r3, down=d3) # db 여부 찾기
    if db == 'db':
        bonus = ws['R29'].value
    else:
        bonus = False
    r4, d4 = get_offset_between_cells('C62', 'AV62')
    fail = get_shifted_cell_value(ws, cell, right=r4, down=d4) # 고장 가능성 찾기
    if isinstance(fail, (int, float)):
        broken = fail
    else:
        if isinstance(fail, str):
            try:
                int(fail)
                broken = fail
            except ValueError:
                broken = None
        else:
            broken = None
    # 판정
    s_c = find_cell_by_value(ws, s) # 기능 찾기

    #####################################
    ### 특수 기능 처리 if 문
    #####################################

    if "AE" in s_c:
        percent = get_shifted_cell_value(ws, s_c, right=14, down=0)
        great = percent // 2
        extreme = percent // 5
    elif "AJ" in s_c:
        r1, d1 = get_offset_between_cells('AJ45', 'AS45')
        percent = get_shifted_cell_value(ws, s_c, right=r1, down=d1)
        great = percent // 2
        extreme = percent // 5
    else:
        r1, d1 = get_offset_between_cells('D40', 'K40')
        percent = get_shifted_cell_value(ws, s_c, right=r1, down=d1)
        great = percent // 2
        extreme = percent // 5

    result, message = CoC_dice(modifier)

    success = ""
    
    if result <= extreme:
        success = "극단적 성공"
    elif result <= great:
        success = "어려운 성공"
    elif result <= percent:
        success = "성공"
    else:
        success = "실패"
    if result == 1:
        success = "대성공"
    if result >= 96:
        if percent < 50:
            success = "대실패"
        else:
            if result == 100:
                success = "대실패"
    if broken:
        if result >= broken:
            success += "+고장"

    script += f"{percent} || 🎲 1d100 = {message} [{success}]"

    r, max_r, rolls = roll_dice_expression(damage)

    
    if success == "극단적 성공" or success == "대성공":
        if tag == '치명타':
            script += f" | {damage} = {rolls} → {r} + {max_r} 치명타!"
        else:
            script += f" | {damage} = {max_r}"
    elif success == "성공" or success == "대단한 성공":
        script += f" | {damage} = {rolls} → {r}"

    if bonus and bonus != "0":
        br, bmax_r, brolls = roll_dice_expression(bonus)
        if success == "극단적 성공" or success == "대성공":
            script += f" | db {bonus} = {bmax_r}"
        elif success == "성공" or success == "대단한 성공":
            script += f" | db {bonus} = {brolls} → {br}"

    return script


def CoC_stat(id, skill, modifier):
    script = ""
    path = f'CoC/{id}.xlsx'

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # 또는 wb["시트이름"]

    cell = find_cell_by_value(ws, skill) # 기능 찾기
    if skill == "교육" or skill == "정신":
        percent = get_shifted_cell_value(ws, cell, right=2, down=0)
    else:
        r1, d1 = get_offset_between_cells('X6', 'Z6')
        percent = get_shifted_cell_value(ws, cell, right=r1, down=d1)
    extreme = percent // 5
    great = percent // 2

    result, message = CoC_dice(modifier)

    success = ""
    
    if result <= extreme:
        success = "극단적 성공"
    elif result <= great:
        success = "어려운 성공"
    elif result <= percent:
        success = "성공"
    else:
        success = "실패"
    if result == 1:
        success = "대성공"
    if result >= 96:
        if percent < 50:
            success = "대실패"
        else:
            if result == 100:
                success = "대실패"

    script += f"{percent} || 🎲 1d100 = {message} [{success}]"
    return script

def CoC_skill(id, skill, modifier):
    script = ""
    path = f'CoC/{id}.xlsx'

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # 또는 wb["시트이름"]

    cell = find_cell_by_value(ws, skill) # 기능 찾기

    #####################################
    ### 특수 기능 처리 if 문
    #####################################

    if "AE" in cell:
        percent = get_shifted_cell_value(ws, cell, right=14, down=0)
        great = percent // 2
        extreme = percent // 5
    elif "AJ" in cell:
        r1, d1 = get_offset_between_cells('AJ45', 'AS45')
        percent = get_shifted_cell_value(ws, cell, right=r1, down=d1)
        great = percent // 2
        extreme = percent // 5
    else:
        r1, d1 = get_offset_between_cells('D40', 'K40')
        percent = get_shifted_cell_value(ws, cell, right=r1, down=d1)
        great = percent // 2
        extreme = percent // 5

    result, message = CoC_dice(modifier)

    success = ""
    
    if result <= extreme:
        success = "극단적 성공"
    elif result <= great:
        success = "어려운 성공"
    elif result <= percent:
        success = "성공"
    else:
        success = "실패"
    if result == 1:
        success = "대성공"
    if result >= 96:
        if percent < 50:
            success = "대실패"
        else:
            if result == 100:
                success = "대실패"

    script += f"{percent} || 🎲 1d100 = {message} [{success}]"
    return script

def CoC_sanity(sanity, modifier):
    percent = sanity
    great = sanity // 2
    extreme = sanity // 5

    result, message = CoC_dice(modifier)

    success = ""
    script = ""
    
    if result <= extreme:
        success = "극단적 성공"
    elif result <= great:
        success = "어려운 성공"
    elif result <= percent:
        success = "성공"
    else:
        success = "실패"
    if result == 1:
        success = "대성공"
    if result >= 96:
        if percent < 50:
            success = "대실패"
        else:
            if result == 100:
                success = "대실패"

    script += f"{percent} || 🎲 1d100 = {message} [{success}]"

    return script


################################################## inSANe



def inSANe_roll(point, modifier = None):

    total, max_possible, rolls = roll_dice_expression("2d6")

    if total == 12:
        success = "크리티컬"
    elif total == 2:
        success = "펌블"
    elif total + int(modifier) >= int(point):
        success = "성공"
    else:
        success = "실패"

    return success, total + int(modifier), rolls


def inSANe_default(id, skill, modifier, ability, fears):
    script = ""
    path = f'inSANe/{id}.xlsx'

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if skill == None:
        a = find_cell_by_value(ws, ability)
        r1, d1 = get_offset_between_cells('L23', 'U23')
        desc = get_shifted_cell_value(ws, a, right=r1, down=d1)
        script += f" ✨ {ability} || {desc}"
        return script

    if skill == '회피':
        plot = int(ability) + 4
        success, total, rolls = inSANe_roll(plot, modifier)
        script += f"{plot} || 🎲 2d6 = {rolls} {"" if modifier == "0" or modifier == 0 else modifier} → {total} [{success}]"
        return script

    cell = find_cell_by_value(ws, skill) # 기능 찾기
    r1, d1 = get_offset_between_cells('S5', 'T5')
    point = get_shifted_cell_value(ws, cell, right=r1, down=d1)

    if ability == "공포판정":
        if skill in fears:
            point = str(int(point) + 2)
        success, total, rolls = inSANe_roll(point, modifier)
        script += f"{point} || 🎲 2d6 = {rolls} {"" if modifier == "0" or modifier == 0 else modifier} → {total} [{success}]"
        return script

    success, total, rolls = inSANe_roll(point, modifier)
    script += f"{point} || 🎲 2d6 = {rolls} {"" if modifier == "0" or modifier == 0 else modifier} → {total} [{success}]"
    if ability:
        a = find_cell_by_value(ws, ability)
        r1, d1 = get_offset_between_cells('L23', 'U23')
        desc = get_shifted_cell_value(ws, a, right=r1, down=d1)
        script += f" ✨ {ability} || {desc}"
    return script

def inSANe_insert_card():
    wb = load_workbook("inSANe/insane.xlsx", data_only=True)
    ws = wb.active

    count = 0
    for cell in ws['A'][1:]:
        if cell.value is not None and str(cell.value).strip() != "":
            count += 1
    
    return count

def inSANe_card(count):
    script = ""
    path = f'inSANe/insane.xlsx'

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    a_values = [cell.value for cell in ws['A'][1:] if cell.value is not None]
    total = len(a_values)
    card = a_values[total - count]
    b_values = [cell.value for cell in ws['B'][1:] if cell.value is not None]
    desc = b_values[total - count]

    script += f"🃏 {card} || {desc}"
    return script

def insane_category(id, category):
    script = ""
    path = f'inSANe/{id}.xlsx'

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    total, max_possible, rolls = roll_dice_expression("2d6")

    category_cols = {
        '폭력': 15,   # O
        '정서': 19,   # S
        '지각': 23,   # W
        '기술': 27,   # AA
        '지식': 31,   # AE
        '괴이': 35    # AI
    }

    if category not in category_cols:
        return "잘못된 분야입니다."

    base_row = 2
    base_col = category_cols[category]
    target_row = base_row + (total - 1)

    skill = ws.cell(row=target_row, column=base_col).value

    script += f"♾️ {category} || {rolls} → {total} [{skill}]"
    return script


################################################## MGLG

def mglg_roll(point):

    total, max_possible, rolls = roll_dice_expression("2d6")

    if total == 12:
        success = "크리티컬"
    elif total == 2:
        success = "펌블"
    elif total >= int(point):
        success = "성공"
    else:
        success = "실패"

    if rolls[0] == rolls[1]:
        success += " | 더블릿 | "
        if rolls[0] == 1: success += "별"
        if rolls[0] == 2: success += "짐승"
        if rolls[0] == 3: success += "힘"
        if rolls[0] == 4: success += "노래"
        if rolls[0] == 5: success += "꿈"
        if rolls[0] == 6: success += "어둠"
        success += " 마소 1점 발생"

    return success, total, rolls

def mglg_category(id, category):
    script = ""
    path = f'MGLG/{id}.xlsx'

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    total, max_possible, rolls = roll_dice_expression("2d6")

    category_cols = {
        '별': 6,   # F
        '짐승': 12,   # L
        '힘': 17,   # Q
        '노래': 23,   # W
        '꿈': 31,   # AE
        '어둠': 37    # AK
    }

    if category not in category_cols:
        return "잘못된 분야입니다."

    base_row = 46
    base_col = category_cols[category]
    target_row = base_row + (total - 1)

    skill = ws.cell(row=target_row, column=base_col).value

    script += f"♾️ {category} || {rolls} → {total} [{skill}]"
    return script


def mglg_table(table):
    script = ""
    path = f'MGLG/table.xlsx'

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[table]
    
    if ws.cell["A7"].value == "" or ws.cell["A7"].value is None or ws.cell["A7"].value is False:
        total, max_possible, rolls = roll_dice_expression("1d6")
    else:
        total, max_possible, rolls = roll_dice_expression("2d6")

    result = ws.cell(row = total, column=2).value

    script += f"🔀 {table} || {rolls} → {total} [{result}]"
    return script


def mglg_default(user, skill):
    script = ""
    path = f'MGLG/{user}.xlsx'

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    cell = find_cell_by_value(ws, skill) # 기능 찾기

    if cell.column == 'F': r1, d1 = get_offset_between_cells('F47', 'I47')
    elif cell.column == 'L': r1, d1 = get_offset_between_cells('L47', 'N47')
    elif cell.column == 'Q': r1, d1 = get_offset_between_cells('Q47', 'S47')
    elif cell.column == 'W':
        if cell.row == 58 or cell.row == '58':
            r1, d1 = get_offset_between_cells('W58', 'AQ58')
        else:
            r1, d1 = get_offset_between_cells('W47', 'AA47')
    elif cell.column == 'AE': r1, d1 = get_offset_between_cells('AE47', 'AH47')
    elif cell.column == 'AK': r1, d1 = get_offset_between_cells('AK47', 'AQ47')

    point = get_shifted_cell_value(ws, cell, right=r1, down=d1)

    success, total, rolls = mglg_roll(point)
    script += f"{point} || 🎲 2d6 = {rolls} → {total} [{success}]"
    
    return script




def m_d66():
    first = random.randint(1, 6)
    second = random.randint(1, 6)
    script = f"🎲 d66 → ({first}, {second})"
    return script